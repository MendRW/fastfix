import pandas as pd
import openpyxl
from openpyxl import load_workbook

def main():
    colfilter = [0, 1, 2]  # Replace with the indices of columns you want to pull
    resolvedby = ["dean.lynch"]  # Names of people who resolved tickets
    df = pd.read_csv('your_file.csv') 

    filtered_df = df[df['ResolvedBy'].isin(resolvedby)]
    result_df = filtered_df.iloc[:, colfilter]

    excel_file = 'your_existing_file.xlsx'

    try:
        # Try to load the existing workbook
        book = load_workbook(excel_file)
    except FileNotFoundError:
        # If file doesn't exist, create a new workbook
        book = openpyxl.Workbook()
        # Remove the default sheet if it exists
        if 'Sheet' in book.sheetnames:
            book.remove(book['Sheet'])
        
    sheet_name = 'Filtered_Data'

    if sheet_name in book.sheetnames:
        book.remove(book[sheet_name])

    book.create_sheet(sheet_name)
    book.save(excel_file)

    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        result_df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"Data has been written to '{sheet_name}' sheet in '{excel_file}'")



if __name__ == "__main__":
    main()